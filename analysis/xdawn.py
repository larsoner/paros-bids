# XDawn-based analysis for PAROS
#
# Todo:
# - 20 control, 16 ASD -> should cull to 32 total
# - Add source space analysis
# - Use DataFrame instead of dict of dicts for latencies and amplitudes

from pathlib import Path
import sys

import numpy as np
import h5io
import mne
import matplotlib.pyplot as plt
import openpyxl

this_dir = Path(__file__).parent
sys.path.append(str(this_dir / '..'))

import paros_bids_config

subjects = [f'sub-{s}' for s in paros_bids_config.subjects]
mbp_path = paros_bids_config.bids_root / 'derivatives' / 'mne-bids-pipeline'

plot_vals = False
save_evoked = True

# 1. Load and process each subject using XDawn
#    1. First as difference to noise (for N100)
#    2. Second as difference between conditions (for N400)


def _get_data(subject, suffix='epo'):
    path = (mbp_path / subject / 'meg' /
            f'{subject}_task-lexicaldecision_{suffix}.fif')
    if suffix == 'epo':
        return mne.read_epochs(path)
    else:
        assert suffix == 'ave'
        return mne.read_evokeds(path)[0]


def _hemi_ch_names(info, hemi):
    if hemi == 'both':
        ch_names = [info['ch_names'][pick]
                    for pick in mne.pick_types(info, meg=True)]
        assert len(ch_names) == 306
    else:
        assert hemi in ('left', 'right')
        ch_names = mne.read_vectorview_selection(
            f'{hemi.capitalize()}-', info=info)
        assert len(ch_names) == 153, len(ch_names)
        assert set(info['ch_names']).issuperset(ch_names)
    return ch_names


# Each cond will be an XDawn transformer (one per hemi type)
peaks = ('N100', 'N400')
conds = ('lexical', 'nonlex')
hemis = ('both', 'left', 'right')
kinds = tuple(
    f'{peak}_{cond}_{hemi}'
    for peak in peaks
    for cond in conds
    for hemi in hemis
)
del peaks, conds, hemis
tlims = {
    kind: (0, 0.3) if kind.split('_')[0] == 'N100' else (0.3, 0.6)
    for kind in kinds
}
extras = {
    kind: 'Xdawn on ' +
    ('evoked vs epochs (all trials)'
     if kind.split('_')[0] == 'N100' else
     'high evoked vs low evoked')
    for kind in kinds
}
titles = {
    kind: (f'{kind} using {"-".join(str(1000 * t) for t in tlims[kind])} '
           f' {extras[kind]}')
    for kind in kinds}


if save_evoked:
    evoked_dir = this_dir / 'evoked'
    evoked_dir.mkdir(exist_ok=True)
    for subject in subjects:
        epochs = _get_data(subject, 'epo').apply_baseline((None, 0))
        assert len(epochs.event_id) == 4
        assert len(epochs) > 200, len(epochs)
        epochs.equalize_event_counts()
        evokeds = list()
        for cond in ('all', 'lexical', 'nonlex'):
            if cond == 'all':
                evokeds.append(epochs.average())
            else:
                evokeds.append(epochs[cond].average())
        mne.write_evokeds(
            evoked_dir / f'{subject}-ave.fif', evokeds, overwrite=True)

xdawn_path = this_dir / 'xdawn_responses.h5'
if not xdawn_path.is_file():
    xdawn_data = {kind: dict() for kind in kinds}
    for si, subject in enumerate(subjects):
        print(
            f'Processing subject {si + 1}/{len(subjects)}: {subject} ')
        epochs = _get_data(subject, 'epo').apply_baseline((None, 0))
        # Eventually we should use the rank, but Xdawn is not built for it yet
        # rank = mne.compute_rank(epochs, tol=1e-6, tol_kind='relative')
        # print(f'rank: {rank}')
        rank = 'full'
        assert len(epochs.event_id) == 4
        assert len(epochs) > 200, len(epochs)
        epochs.equalize_event_counts()
        assert len(epochs) > 160, len(epochs)
        assert set(epochs.event_id) == {
            'lexical/high', 'lexical/low', 'nonlex/high', 'nonlex/low'}, \
                epochs.event_id
        for kind in kinds:
            peak, cond, hemi = kind.split('_')
            want_ids = [f'{cond}/{lohi}' for lohi in ('high', 'low')]
            assert kind.count('_') == 2  # 3 parts
            reg = 0.1
            method = 'shrunk'
            method_params = dict(shrunk=dict(shrinkage=[reg]))
            fit_epochs = epochs[cond]
            assert len(fit_epochs.event_id) == 2
            assert set(fit_epochs.event_id) == set(want_ids)
            if peak == 'N100':
                key = 'auditory'
                fit_epochs = mne.epochs.combine_event_ids(
                    epochs, list(epochs.event_id), {key: 1000})
                fit_epochs.crop(0, 0.3)
            else:
                assert peak == 'N400'
                key = f'{cond}/low'
                fit_epochs.crop(0.3, 0.6)
            # Restrict to proper channels
            ch_names = _hemi_ch_names(epochs.info, hemi)
            fit_epochs.pick(ch_names)
            epochs_picked = epochs[cond].pick(ch_names)
            if peak == 'N100':
                signal_cov_epo = fit_epochs
            else:
                # Bias cov is the cond/low evoked cov
                signal_cov_epo = mne.EpochsArray(
                    fit_epochs[f'{cond}/high'].average().data[np.newaxis],
                    fit_epochs.info, tmin=fit_epochs.tmin)
            signal_cov = mne.compute_covariance(
                signal_cov_epo, method=method, method_params=method_params,
                rank=rank, verbose='error')  # Xdawn limitation
            xd = mne.preprocessing.Xdawn(
                n_components=1, correct_overlap=False, reg=reg,
                signal_cov=signal_cov)
            xd.fit(fit_epochs[key])
            if peak == 'N100':
                tc = xd.transform(epochs_picked.average())
            else:
                assert peak == 'N400'
                tc = (xd.transform(epochs_picked[f'{cond}/low'].average()) -
                      xd.transform(epochs_picked[f'{cond}/high'].average()))
            assert tc.shape == (1, len(epochs.times))
            assert xd.patterns_[key].shape == (len(ch_names), len(ch_names))
            tc = tc[0]
            pattern = xd.patterns_[key][0]
            times = epochs.times
            assert tc.shape == (len(epochs.times),)
            assert pattern.shape == (len(ch_names),)
            xdawn_data[kind][subject] = dict(
                tc=tc, pattern=pattern, times=times)
            # evoked = epochs.average()
            # evoked.data = pattern[:, np.newaxis] @ tc[np.newaxis]
            # evoked.plot()
            del tc, pattern
    h5io.write_hdf5(xdawn_path, xdawn_data)

xdawn_data = h5io.read_hdf5(xdawn_path)

# 2. Align all subjects waveforms to the first PC for both types
flips = dict()
for kind in kinds:
    this_data = np.array([v['tc'] for v in xdawn_data[kind].values()], float)
    _, s, v = np.linalg.svd(this_data, full_matrices=False)
    times = xdawn_data[kind][subjects[0]]['times']
    exp_var = np.cumsum(s ** 2)
    exp_var = 100 * exp_var[0] / exp_var[-1]
    print(f'Alignment of {kind.ljust(20)} exp var: {exp_var:0.1f}%')
    pc = v[0]
    flips[kind] = np.sign(this_data @ v[0])
    assert not (flips[kind] == 0).any()
    # now align the waveforms to be negative on average in the window of
    # interest
    this_data *= flips[kind][:, np.newaxis]
    t_mask = (times >= tlims[kind][0]) & (times <= tlims[kind][1])
    this_data = this_data[:, t_mask]
    if this_data.mean() > 0:
        flips[kind] *= -1

# 3. Plot big matrix of subject waveforms and topomaps for each type
# 4. Extract peak from each subject and each condition (latency, amplitude)
measures = ['latency', 'amplitude']
csv_header = ['subject']
csv_values = list()
latencies = dict()
amplitudes = dict()
for ki, kind in enumerate(kinds):
    n_col = int(np.ceil(np.sqrt(len(subjects))))
    n_row = (len(subjects) - 1) // n_col + 1
    csv_header.extend(f'{kind}_{meas}' for meas in measures)
    if plot_vals:
        fig, axes = plt.subplots(
            n_row, n_col, figsize=(n_col * 2.5, n_row * 1.5),
            constrained_layout=True, squeeze=False)
        fig.suptitle(titles[kind], fontsize=8)
    xlim = [-0.2, 1.2]
    xticks = np.array([0, 0.4, 0.8, 1.2])
    latencies[kind] = dict()
    amplitudes[kind] = dict()
    for si, subject in enumerate(subjects):
        if ki == 0:
            csv_values.append([subject])
        this_data = xdawn_data[kind][subject]
        ri, ci = divmod(si, n_col)
        fl = flips[kind][si]
        d = this_data['tc'] * fl * 1e12  # to pT
        times = this_data['times']
        mask = (times >= tlims[kind][0]) & (times <= tlims[kind][1])
        idx = np.argmin(d * mask)
        latencies[kind][subject] = lat = times[idx]
        amplitudes[kind][subject] = amp = d[idx]
        # prepare CSV
        for measure in measures:
            if measure == 'latency':
                val = f'{1e3 * latencies[kind][subject]:0.1f}'  # s->ms
            else:
                assert measure == 'amplitude'
                val = f'{1e3 * amplitudes[kind][subject]:0.3f}'  # pT->fT
            csv_values[si].append(val)
        if not plot_vals:
            continue
        ax = axes[ri, ci]
        ax.plot(times, d, lw=1, zorder=3)
        ax.plot(lat, amp,
                'ro', ms=5, markerfacecolor='none', zorder=4)
        ax.text(lat, amp, f'  {1e3 * lat:0.1f}: {1e3 * amp:0.0f}',
                ha='left', va='center', fontsize=6, color='r', zorder=4)
        ax.set_title(subject.lstrip('sub-'), fontsize=6)
        ax.set_xticks(xticks)
        ax.set_xlim(xticks[[0, -1]])
        # inset axis
        ax = ax.inset_axes([0.7, 0.3, 0.3, 0.7])
        info = _get_data(subject, 'ave').info
        hemi = kind.split('_')[2]
        picks = np.where(
            np.in1d(info['ch_names'], _hemi_ch_names(info, hemi)))[0]
        assert len(picks) in (306, 153)
        info = mne.pick_info(info, picks)
        assert len(picks) == len(this_data['pattern'])
        picks = mne.pick_types(info, meg='mag')
        info = mne.pick_info(info, picks)
        mne.viz.plot_topomap(
            fl * this_data['pattern'][picks], info, axes=ax, sensors=False,
            contours=0)
        ax.axis('off')
        del info
    if plot_vals:
        fig.savefig(this_dir / f'resp_{kind}.png')

# 5. Add matching information
static_dir = this_dir / '..' / 'static'
wb = openpyxl.load_workbook(
    static_dir / 'GABA_subject_information.xlsx')
ws = [ws for ws in wb.worksheets if ws.title == 'Matches'][0]
asd_col, con_col = 1, 4
assert ws.cell(1, asd_col).value == 'ASD', ws.cell(1, asd_col).value
assert ws.cell(1, con_col).value == 'Control', ws.cell(1, con_col).value
asd = list()
con = list()
for ri in range(2, 100):
    val = ws.cell(ri, asd_col).value
    if not val:
        break
    asd.append('sub-' + val.split('_')[-1])
    val = ws.cell(ri, con_col).value
    con.append('sub-' + val.split('_')[-1])
assert set(asd).intersection(set(con)) == set()
missing_match = set(subjects).difference(set(asd).union(set(con)))
if missing_match:
    print(f'Missing from matching map: {sorted(missing_match)}')
csv_header.append('ASD_Control_Map')
for si, subject in enumerate(subjects):
    try:
        ii = asd.index(subject) + 1
    except ValueError:
        try:
            ii = -(con.index(subject) + 1)
        except ValueError:
            assert subject in missing_match
            ii = 0
    else:
        assert subject not in con  # in exactly one list
    csv_values[si].append(str(ii))

# 6. Add MRIS information
wb = openpyxl.load_workbook(
    static_dir / 'all_nbwr_mrs_results_csfcorr_fits_20180417.xlsx')
ws = [ws for ws in wb.worksheets if ws.title == 'FSLcorr_metab'][0]
assert ws.cell(2, 1).value == 'subject'
n_col = 0
for ci in range(2, 100):
    hdr = ws.cell(2, ci).value
    if not hdr:
        break
    csv_header.append(hdr)
    n_col += 1
used = np.zeros(len(subjects), bool)
extra_mris = set()
for ri in range(3, 100):
    subject = ws.cell(ri, 1).value
    if not subject:
        break
    subject = f'sub-{subject.split("nbwr")[-1]}'
    try:
        si = subjects.index(subject)
    except ValueError:
        extra_mris.add(subject)
        continue
    used[si] = True
    csv_values[si].extend(
        str(float(ws.cell(ri, 2 + ii).value)) for ii in range(n_col))
missing_mris = set()
for si in np.where(~used)[0]:
    csv_values[si].extend([''] * n_col)
    missing_mris.add(subjects[si])
if missing_mris:
    print(f'Missing from MRIS data:    {sorted(missing_mris)}')
if extra_mris:
    print(f'In MRIS but not MEG data:  {sorted(extra_mris)}')

# 7. Write to CSV
csv_values = np.array(csv_values, dtype=str)
assert csv_values.shape == (len(subjects), len(csv_header)), csv_values.shape
with open(this_dir / 'xdawn_measures.csv', 'w') as fid:
    fid.write(','.join(csv_header) + '\n')
    for row in csv_values:
        fid.write(','.join(row) + '\n')

# 6. Project XDawn components to source space for visualization
