from pathlib import Path
import sys
import openpyxl

this_dir = Path(__file__).parent
if str(this_dir) not in sys.path:
    sys.path.insert(0, str(this_dir))

import paros_bids_config  # noqa


def get_subjects():
    return [f'sub-{s}' for s in paros_bids_config.subjects]


def get_groups():
    subjects = get_subjects()
    static_dir = this_dir / 'static'
    wb = openpyxl.load_workbook(
        static_dir / 'GABA_subject_information.xlsx')
    ws = [ws for ws in wb.worksheets if ws.title == 'Matches'][0]
    asd_col, con_col = 1, 4
    assert ws.cell(1, asd_col).value == 'ASD', ws.cell(1, asd_col).value
    assert ws.cell(1, con_col).value == 'Control', ws.cell(1, con_col).value
    asd = list()
    con = list()
    for ri in range(2, 100):
        val = ws.cell(ri, asd_col).value
        if not val:
            break
        asd.append('sub-' + val.split('_')[-1])
        val = ws.cell(ri, con_col).value
        con.append('sub-' + val.split('_')[-1])
    assert set(asd).intersection(set(con)) == set()
    missing_match = set(subjects).difference(set(asd).union(set(con)))
    if missing_match:
        print(f'Missing from matching map: {sorted(missing_match)}')
    missing_con = set(con).difference(set(subjects))
    if missing_con:
        print(f'Missing from control data: {sorted(missing_con)}')
    # 421 is missing its match
    print('  Removing 451 from con')
    con.pop(con.index('sub-451'))
    missing_asd = set(asd).difference(set(subjects))
    if missing_asd:
        print(f'Missing from asd data:     {sorted(missing_asd)}')
        for key in missing_asd:
            print(f'  Removing {key} from asd')
            asd.pop(asd.index(key))

    assert len(subjects) == 36, len(subjects)
    groups = {
        'grand-average': asd + con,
        'asd': asd,
        'control': con,
    }
    assert len(asd) == 16, len(asd)
    assert len(con) == 16, len(con)
    return groups
